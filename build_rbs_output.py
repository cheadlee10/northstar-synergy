import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

OUT = r"C:\Users\chead\OneDrive\Documents\RBS_Savings_Analysis.xlsx"

facilities = [
    {"facility":"Dallas",      "rbs_cpp":2.2894,"stop_cpp":2.5677,"cpp_diff":0.2783,"total_savings":1598808,"daily_avg_vol":27866,"forecast_ppr":150.5,"labor_ppd":240.5,"gas_maint_ppm":0.3119,"avg_miles":100.8},
    {"facility":"Atlanta",     "rbs_cpp":2.7449,"stop_cpp":3.0666,"cpp_diff":0.3217,"total_savings":1444734,"daily_avg_vol":21779,"forecast_ppr":124.9,"labor_ppd":238.4,"gas_maint_ppm":0.3321,"avg_miles":109.5},
    {"facility":"Northern NJ", "rbs_cpp":1.9651,"stop_cpp":2.1371,"cpp_diff":0.1720,"total_savings":1151791,"daily_avg_vol":32489,"forecast_ppr":193.1,"labor_ppd":279.6,"gas_maint_ppm":0.3421,"avg_miles":57.2},
    {"facility":"Pittsburgh",  "rbs_cpp":2.1221,"stop_cpp":2.6804,"cpp_diff":0.5583,"total_savings":1143651,"daily_avg_vol":9935, "forecast_ppr":159.8,"labor_ppd":234.7,"gas_maint_ppm":0.3421,"avg_miles":87.8},
    {"facility":"Petaluma",    "rbs_cpp":2.2898,"stop_cpp":2.8387,"cpp_diff":0.5490,"total_savings":1050173,"daily_avg_vol":9278, "forecast_ppr":173.2,"labor_ppd":280.8,"gas_maint_ppm":0.4392,"avg_miles":88.9},
    {"facility":"Boise",       "rbs_cpp":2.0765,"stop_cpp":2.7071,"cpp_diff":0.6306,"total_savings":1007976,"daily_avg_vol":7753, "forecast_ppr":162.6,"labor_ppd":233.0,"gas_maint_ppm":0.3476,"avg_miles":86.0},
    {"facility":"Detroit",     "rbs_cpp":2.3432,"stop_cpp":2.6660,"cpp_diff":0.3228,"total_savings":867083, "daily_avg_vol":13027,"forecast_ppr":154.2,"labor_ppd":257.2,"gas_maint_ppm":0.3350,"avg_miles":91.9},
    {"facility":"Denver",      "rbs_cpp":2.1560,"stop_cpp":2.3277,"cpp_diff":0.1718,"total_savings":818971, "daily_avg_vol":23127,"forecast_ppr":171.8,"labor_ppd":268.8,"gas_maint_ppm":0.3476,"avg_miles":72.5},
    {"facility":"Toledo",      "rbs_cpp":2.3359,"stop_cpp":3.4872,"cpp_diff":1.1514,"total_savings":810087, "daily_avg_vol":3413, "forecast_ppr":150.0,"labor_ppd":236.2,"gas_maint_ppm":0.3344,"avg_miles":124.7},
    {"facility":"Tucson",      "rbs_cpp":1.9905,"stop_cpp":2.5824,"cpp_diff":0.5919,"total_savings":673984, "daily_avg_vol":5523, "forecast_ppr":173.0,"labor_ppd":235.6,"gas_maint_ppm":0.3926,"avg_miles":81.9},
    {"facility":"Miami",       "rbs_cpp":2.0036,"stop_cpp":2.1278,"cpp_diff":0.1242,"total_savings":483629, "daily_avg_vol":18887,"forecast_ppr":171.1,"labor_ppd":243.1,"gas_maint_ppm":0.3321,"avg_miles":70.5},
    {"facility":"Buckeye",     "rbs_cpp":2.1461,"stop_cpp":2.2317,"cpp_diff":0.0856,"total_savings":302364, "daily_avg_vol":17131,"forecast_ppr":166.6,"labor_ppd":248.2,"gas_maint_ppm":0.3926,"avg_miles":86.4},
]

total_savings = sum(f['total_savings'] for f in facilities)
avg_cpp_diff  = sum(f['cpp_diff'] for f in facilities) / len(facilities)
total_vol     = sum(f['daily_avg_vol'] for f in facilities)

DARK_BLUE  = "1F3864"
MID_BLUE   = "2F5597"
GREEN_POS  = "E2EFDA"
GREEN_TXT  = "375623"
GREY_ROW   = "F2F2F2"
YELLOW_HL  = "FFFF00"

thin = Side(style='thin', color="BFBFBF")
bdr  = Border(left=thin, right=thin, bottom=thin, top=thin)

def hdr(cell, bg=DARK_BLUE):
    cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = bdr

wb = openpyxl.Workbook()

# ── SHEET 1: EXECUTIVE SUMMARY ────────────────────────────────────────────────
ws = wb.active
ws.title = "Executive Summary"
ws.sheet_view.showGridLines = False

for col, w in zip('ABCDEFG', [2,28,16,16,16,16,16]):
    ws.column_dimensions[col].width = w

ws.merge_cells('B1:G1')
ws['B1'] = "RBS Pilot — Savings Analysis  |  12 Highlighted Facilities"
ws['B1'].font = Font(name="Calibri", bold=True, size=15, color=DARK_BLUE)
ws['B1'].alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 30

ws.merge_cells('B2:G2')
ws['B2'] = "Route-Based Settlement vs. Stop Payment  |  2026 Annual Projection  |  February 23, 2026"
ws['B2'].font = Font(name="Calibri", size=10, color="595959", italic=True)
ws['B2'].alignment = Alignment(horizontal="left")
ws.row_dimensions[2].height = 16

ws.row_dimensions[3].height = 6

# KPI row
kpis = [
    ("Total Annual Savings", f"${total_savings:,.0f}", "12 Pilot Facilities"),
    ("Avg CPP Improvement",  f"${avg_cpp_diff:.4f}",   "Per Package"),
    ("Largest CPP Gap",      "Toledo  $1.1514",         "3,413 daily pkgs"),
    ("Top $ Savings",        "Dallas  $1,598,808",      "27,866 daily pkgs"),
    ("Total Daily Volume",   f"{total_vol:,.0f}",       "Across 12 Facilities"),
]
for i, (label, val, sub) in enumerate(kpis):
    col = get_column_letter(i+2)
    for r, txt, sz, bold in [(4,'',9,False),(5,val,13,True),(6,label,9,True),(7,sub,9,False)]:
        c = ws[f'{col}{r}']
        c.value = txt
        c.font = Font(name="Calibri", bold=bold, size=sz, color="FFFFFF" if r<7 else "595959")
        c.fill = PatternFill("solid", fgColor=MID_BLUE if r<7 else "EBF3FB")
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 6
    ws.row_dimensions[5].height = 26
    ws.row_dimensions[6].height = 18
    ws.row_dimensions[7].height = 16

ws.row_dimensions[8].height = 8

# Table header
hdrs = ["Facility","RBS CPP","Stop Pmt CPP","CPP Savings/Pkg","Annual Savings","Daily Avg Vol"]
for i, h in enumerate(hdrs):
    hdr(ws[f'{get_column_letter(i+2)}9'], bg=DARK_BLUE)
    ws[f'{get_column_letter(i+2)}9'].value = h
ws.row_dimensions[9].height = 26

for r, fac in enumerate(facilities, 10):
    bg = GREY_ROW if r % 2 == 0 else "FFFFFF"
    row_data = [
        (fac['facility'],          "left",   None,       True,  DARK_BLUE),
        (f"${fac['rbs_cpp']:.4f}", "center", None,       False, None),
        (f"${fac['stop_cpp']:.4f}","center", None,       False, None),
        (fac['cpp_diff'],          "center", '$0.0000',  True,  GREEN_TXT),
        (fac['total_savings'],     "right",  '$#,##0',   True,  GREEN_TXT),
        (fac['daily_avg_vol'],     "right",  '#,##0',    False, None),
    ]
    for i, (val, align, fmt, bold_, color_) in enumerate(row_data):
        c = ws[f'{get_column_letter(i+2)}{r}']
        c.value = val
        c.border = bdr
        c.alignment = Alignment(horizontal=align, vertical="center", indent=1 if align=="left" else 0)
        cfg = GREEN_POS if i in [3,4] else bg
        c.fill = PatternFill("solid", fgColor=cfg)
        c.font = Font(name="Calibri", bold=bold_, size=10, color=color_ or "000000")
        if fmt:
            c.number_format = fmt
    ws.row_dimensions[r].height = 18

# Totals
r = 10 + len(facilities)
ws.row_dimensions[r].height = 20
totals = [("TOTAL / AVERAGE","left",None,True,"FFFFFF"),
          ("","center",None,False,"FFFFFF"),
          ("","center",None,False,"FFFFFF"),
          (f"${avg_cpp_diff:.4f} avg","center",None,True,"FFFFFF"),
          (total_savings,"right","$#,##0",True,"FFFFFF"),
          (total_vol,"right","#,##0",True,"FFFFFF")]
for i,(val,align,fmt,bold_,color_) in enumerate(totals):
    c = ws[f'{get_column_letter(i+2)}{r}']
    c.value = val
    c.font = Font(name="Calibri", bold=bold_, size=10, color=color_)
    c.fill = PatternFill("solid", fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal=align, vertical="center", indent=1 if align=="left" else 0)
    c.border = bdr
    if fmt: c.number_format = fmt

# ── SHEET 2: SAVINGS DRIVERS ──────────────────────────────────────────────────
ws2 = wb.create_sheet("Savings Drivers")
ws2.sheet_view.showGridLines = False
for col, w in zip('ABCDEFGHI', [2,20,14,14,14,14,14,14,22]):
    ws2.column_dimensions[col].width = w

ws2.merge_cells('B1:I1')
ws2['B1'] = "What Drives the Savings — Component Analysis"
ws2['B1'].font = Font(name="Calibri", bold=True, size=14, color=DARK_BLUE)
ws2['B1'].alignment = Alignment(horizontal="left", vertical="center")
ws2.row_dimensions[1].height = 28

ws2.merge_cells('B2:I2')
ws2['B2'] = "Three components explain the CPP gap: (1) Route Density — fewer packages/route = stop payment becomes expensive per unit; (2) Route Length — longer miles amplify gas/maint cost under stop payment; (3) Stop Fee Rate — high rate at low-density sites stacks both effects."
ws2['B2'].font = Font(name="Calibri", size=10, color="595959", italic=True)
ws2['B2'].alignment = Alignment(horizontal="left", wrap_text=True)
ws2.row_dimensions[2].height = 36
ws2.row_dimensions[3].height = 8

drv_hdrs = ["Facility","Pcs/Route","Miles/Route","Gas+Maint/Mi","Labor/Day","CPP Gap","Annual Savings","Key Driver"]
for i, h in enumerate(drv_hdrs):
    hdr(ws2[f'{get_column_letter(i+2)}4'], bg=MID_BLUE)
    ws2[f'{get_column_letter(i+2)}4'].value = h
ws2.row_dimensions[4].height = 26

ppr_vals   = [f['forecast_ppr'] for f in facilities]
miles_vals = [f['avg_miles']    for f in facilities]
ppr_min, ppr_max     = min(ppr_vals), max(ppr_vals)
miles_min, miles_max = min(miles_vals), max(miles_vals)

def key_driver(fac):
    if fac['cpp_diff'] > 1.0:   return "Very low density + high stop fee"
    if fac['forecast_ppr'] < 130: return "Low density (pcs/route)"
    if fac['avg_miles'] > 100:   return "Long routes (miles)"
    if fac['cpp_diff'] > 0.5:    return "Low density + long routes"
    if fac['daily_avg_vol'] > 20000: return "High volume — labor containment"
    return "Moderate density + stop fee rate"

for r, fac in enumerate(facilities, 5):
    bg = GREY_ROW if r % 2 == 0 else "FFFFFF"
    # PPR heat: low PPR = more orange (bigger savings driver)
    ppr_norm = (ppr_max - fac['forecast_ppr']) / (ppr_max - ppr_min) if ppr_max != ppr_min else 0
    ppr_r = int(255); ppr_g = int(255 * (1 - ppr_norm * 0.7)); ppr_b = int(255 * (1 - ppr_norm))
    ppr_bg = f"{min(255,ppr_r):02X}{min(255,ppr_g):02X}{min(255,ppr_b):02X}"

    miles_norm = (fac['avg_miles'] - miles_min) / (miles_max - miles_min) if miles_max != miles_min else 0
    mi_r = int(255); mi_g = int(255 * (1 - miles_norm * 0.6)); mi_b = int(255 * (1 - miles_norm))
    mi_bg = f"{min(255,mi_r):02X}{min(255,mi_g):02X}{min(255,mi_b):02X}"

    row_vals = [
        (fac['facility'],          bg,        "left",   None,       True,  DARK_BLUE),
        (fac['forecast_ppr'],      ppr_bg,    "center", "0.0",      False, "000000"),
        (fac['avg_miles'],         mi_bg,     "center", "0.0",      False, "000000"),
        (fac['gas_maint_ppm'],     bg,        "center", "$0.0000",  False, "000000"),
        (fac['labor_ppd'],         bg,        "center", "$#,##0.00",False, "000000"),
        (fac['cpp_diff'],          GREEN_POS, "center", "$0.0000",  True,  GREEN_TXT),
        (fac['total_savings'],     GREEN_POS, "right",  "$#,##0",   True,  GREEN_TXT),
        (key_driver(fac),          "EBF3FB",  "left",   None,       False, "1F3864"),
    ]
    for i, (val, bg_, align, fmt, bold_, color_) in enumerate(row_vals):
        c = ws2[f'{get_column_letter(i+2)}{r}']
        c.value = val
        c.border = bdr
        c.fill = PatternFill("solid", fgColor=bg_)
        c.font = Font(name="Calibri", bold=bold_, size=10, color=color_)
        c.alignment = Alignment(horizontal=align, vertical="center", indent=1 if align=="left" else 0, wrap_text=(i==7))
        if fmt: c.number_format = fmt
    ws2.row_dimensions[r].height = 18

# Findings box
r_n = 5 + len(facilities) + 1
ws2.merge_cells(f'B{r_n}:I{r_n+4}')
ws2[f'B{r_n}'] = (
    "KEY FINDINGS:\n"
    "1. TOLEDO: Highest CPP gap ($1.15) — stop fee rate is very high relative to volume (3,413 daily pkgs). "
    "RBS fixes the route cost regardless of package count — massive per-unit saving.\n"
    "2. BOISE & PITTSBURGH ($0.63, $0.56): Low density + moderate miles. Stop payment compounds inefficiency "
    "at each stop; RBS pays once per route.\n"
    "3. DALLAS & NORTHERN NJ: Large volume facilities. Smaller CPP gap ($0.17–0.28) but big absolute savings "
    "because volume is high (27K and 32K daily packages).\n"
    "4. BUCKEYE & MIAMI: Smallest gaps ($0.09, $0.12) — relatively dense routes. Savings come from "
    "labor/operational efficiency rather than density."
)
ws2[f'B{r_n}'].font = Font(name="Calibri", size=10, color=DARK_BLUE)
ws2[f'B{r_n}'].fill = PatternFill("solid", fgColor="EBF3FB")
ws2[f'B{r_n}'].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
for rr in range(r_n, r_n+5):
    ws2.row_dimensions[rr].height = 20

# ── SHEET 3: CHART DATA ────────────────────────────────────────────────────────
ws3 = wb.create_sheet("Charts")
ws3.sheet_view.showGridLines = False
for col, w in zip('ABCD', [2,22,18,16]):
    ws3.column_dimensions[col].width = w

for h, col in [("Facility","B"),("Annual Savings","C"),("CPP Gap","D")]:
    hdr(ws3[f'{col}1'], bg=MID_BLUE)
    ws3[f'{col}1'].value = h

for r, fac in enumerate(facilities, 2):
    ws3[f'B{r}'] = fac['facility']
    ws3[f'C{r}'] = fac['total_savings']
    ws3[f'D{r}'] = fac['cpp_diff']
    ws3[f'C{r}'].number_format = '$#,##0'
    ws3[f'D{r}'].number_format = '$0.0000'

n = len(facilities)

chart1 = BarChart()
chart1.type = "bar"
chart1.title = "Annual Savings by Facility"
chart1.y_axis.title = "Annual Savings ($)"
chart1.style = 10
chart1.width = 24
chart1.height = 14
chart1.add_data(Reference(ws3, min_col=3, max_col=3, min_row=1, max_row=n+1), titles_from_data=True)
chart1.set_categories(Reference(ws3, min_col=2, min_row=2, max_row=n+1))
chart1.series[0].graphicalProperties.solidFill = MID_BLUE
ws3.add_chart(chart1, "F2")

chart2 = BarChart()
chart2.type = "bar"
chart2.title = "CPP Gap by Facility  (Stop Payment − RBS)"
chart2.y_axis.title = "CPP Savings per Package ($)"
chart2.style = 10
chart2.width = 24
chart2.height = 14
chart2.add_data(Reference(ws3, min_col=4, max_col=4, min_row=1, max_row=n+1), titles_from_data=True)
chart2.set_categories(Reference(ws3, min_col=2, min_row=2, max_row=n+1))
chart2.series[0].graphicalProperties.solidFill = "C55A11"
ws3.add_chart(chart2, "F34")

wb.save(OUT)
print(f"SAVED: {OUT}")
print(f"Total savings: ${total_savings:,.0f}")
print(f"Facilities: {len(facilities)}")
