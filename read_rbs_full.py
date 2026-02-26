import openpyxl
import sys
from collections import defaultdict

FILE = r"C:\Users\chead\.openclaw\workspace\RBS_unzipped\RBS Budget Impact CCH.xlsx"
wb = openpyxl.load_workbook(FILE, read_only=True, data_only=True)

# === SUMMARY - full read ===
print("=== SUMMARY (FULL) ===", flush=True)
ws = wb["Summary"]
for row in ws.iter_rows(values_only=True):
    if any(v is not None for v in row):
        print([str(v)[:40] if v is not None else '' for v in row], flush=True)

# === LOUISVILLE TAB ===
print("\n=== LOUISVILLE ===", flush=True)
ws = wb["Louisville"]
for i, row in enumerate(ws.iter_rows(max_row=50, values_only=True), 1):
    if any(v is not None for v in row):
        print(f"Row {i}: {[str(v)[:30] if v is not None else '' for v in row[:20]]}", flush=True)

# === BY FACILITY - aggregate totals ===
print("\n=== BY FACILITY - ALL ROWS ===", flush=True)
ws = wb["By Facility"]
headers = None
fac_data = []
for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    if i == 4:  # header row
        headers = [str(v) if v is not None else '' for v in row]
        print(f"Headers: {headers[:20]}", flush=True)
    elif i > 4 and any(v is not None for v in row):
        fac_data.append(row)

print(f"Facility rows: {len(fac_data)}", flush=True)

# Sort by projected impact if available
# Col structure: Branch, 2025 Adhoc, Current Base, Current CPS, [monthly stop counts...]
# Let's compute total stops and show branch, adhoc, base, cps
print("\nTop facilities by Current Base (descending):", flush=True)
fac_sorted = sorted(fac_data, key=lambda r: (r[3] if r[3] is not None else 0), reverse=True)
for row in fac_sorted[:25]:
    branch = row[1] if row[1] is not None else ''
    adhoc = row[2] if row[2] is not None else 0
    base = row[3] if row[3] is not None else 0
    cps = row[4] if row[4] is not None else 0
    print(f"  {str(branch):<20} Adhoc%={float(adhoc)*100:.2f}%  Base={float(base):>10,.0f}  CPS={float(cps):.4f}", flush=True)

# === BY TERRITORY - find totals ===
print("\n=== BY TERRITORY - Aggregates by Branch ===", flush=True)
ws = wb["By Territory"]
branch_data = defaultdict(lambda: {'territories': 0, 'impact': 0, 'current_base': 0, 'new_base': 0})
for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    if i <= 3:
        continue
    if row[0] is None:
        continue
    branch = str(row[0])
    if branch.lower() in ['branch', 'sample branch', 'total', '']:
        continue
    try:
        branch_data[branch]['territories'] += 1
        branch_data[branch]['impact'] += float(row[11]) if row[11] is not None else 0
        branch_data[branch]['current_base'] += float(row[6]) if row[6] is not None else 0
        branch_data[branch]['new_base'] += float(row[8]) if row[8] is not None else 0
    except:
        pass

# Sort by impact desc
sorted_branches = sorted(branch_data.items(), key=lambda x: x[1]['impact'], reverse=True)
print(f"{'Branch':<25} {'Territories':>12} {'In-Yr Impact':>15} {'Curr Base':>12} {'New Base':>12} {'Savings':>12}", flush=True)
print("-"*90, flush=True)
total_impact = 0
total_territories = 0
for branch, d in sorted_branches[:30]:
    savings = d['current_base'] - d['new_base']
    print(f"{branch:<25} {d['territories']:>12,} {d['impact']:>15,.0f} {d['current_base']:>12,.0f} {d['new_base']:>12,.0f} {savings:>12,.0f}", flush=True)
    total_impact += d['impact']
    total_territories += d['territories']

print(f"\n{'TOTAL':<25} {total_territories:>12,} {total_impact:>15,.0f}", flush=True)

wb.close()
print("\nDONE", flush=True)
