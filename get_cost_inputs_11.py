import openpyxl, json

FILE = r"C:\Users\chead\.openclaw\workspace\RBS_unzipped\RBS Budget Impact CCH.xlsx"
wb = openpyxl.load_workbook(FILE, read_only=True, data_only=True)

# Top 11 highlighted facilities (by Total Savings desc)
top11 = ['Sacramento','Columbus','Dallas','Atlanta','JFK','Queens',
         'Northern NJ','Pittsburgh','Petaluma','Boise','Detroit']

# By Facility data - RBS, Stop Payment, CPP diff, Total Savings, Daily Avg
ws_fac = wb["By Facility"]
fac_data = {}
for i, row in enumerate(ws_fac.iter_rows(values_only=True), 1):
    if i > 4 and row[1] is not None:
        branch = str(row[1])
        if branch in top11:
            fac_data[branch] = {
                'current_base': row[3],
                'current_cps': row[4],
                'daily_avg': row[79],      # col 80 = daily avg routes
                'rbs_cpp': row[81],        # CD
                'stop_cpp': row[82],       # CE
                'cpp_diff': row[83],       # CF
                'total_savings': row[84],  # CG
                'daily_avg_vol': row[85],  # CH = Daily Average
                'cpp_impact_pct': row[86], # CI = CPP Impact %
            }

# Cost Inputs
ws_ci = wb["Cost Inputs"]
cost_data = {}
for i, row in enumerate(ws_ci.iter_rows(values_only=True), 1):
    if i > 1 and row[0] is not None:
        fac = str(row[0])
        if fac in top11:
            cost_data[fac] = {
                'forecast_ppr': row[1],
                'gas_maint_ppm': row[2],
                'labor_ppd': row[3],
                'avg_miles': row[4]
            }

wb.close()

output = []
for fac in top11:
    fd = fac_data.get(fac, {})
    ci = cost_data.get(fac, {})
    row = {'facility': fac}
    row.update(fd)
    row.update(ci)
    output.append(row)

print(json.dumps(output, indent=2, default=str))
