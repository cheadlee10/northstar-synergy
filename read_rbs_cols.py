import openpyxl
import sys

FILE = r"C:\Users\chead\.openclaw\workspace\RBS_unzipped\RBS Budget Impact CCH.xlsx"

# First pass: read_only to get data from By Facility tab, columns A-CG (1-85)
wb = openpyxl.load_workbook(FILE, read_only=True, data_only=True)
ws = wb["By Facility"]

print("=== BY FACILITY - Columns A-B + CD-CG (cols 1-2, 82-85) ===", flush=True)
header_row = None
all_rows = []
for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    if i == 4:
        header_row = list(row)
        # Print headers around CD-CG
        print(f"Col 80-87 headers: {header_row[79:87]}", flush=True)
        print(f"Col 1-5 headers: {header_row[:5]}", flush=True)
    elif i > 4:
        if any(v is not None for v in row):
            # Store: row index, branch (col B = index 1), CD=81, CE=82, CF=83, CG=84
            all_rows.append({
                'row_idx': i,
                'branch': row[1] if len(row) > 1 else None,
                'current_base': row[3] if len(row) > 3 else None,
                'current_cps': row[4] if len(row) > 4 else None,
                # Try various column ranges for CD-CG
                'col80': row[79] if len(row) > 79 else None,
                'col81': row[80] if len(row) > 80 else None,
                'col82': row[81] if len(row) > 81 else None,
                'col83': row[82] if len(row) > 82 else None,
                'col84': row[83] if len(row) > 83 else None,
                'col85': row[84] if len(row) > 84 else None,
                'col86': row[85] if len(row) > 85 else None,
                'col87': row[86] if len(row) > 86 else None,
            })

wb.close()

print(f"\nTotal data rows: {len(all_rows)}", flush=True)
print(f"\n{'Branch':<25} {'Base':>10} {'CPS':>8} {'C80':>12} {'C81':>12} {'C82':>12} {'C83':>12} {'C84':>12} {'C85':>12}", flush=True)
print("-"*110, flush=True)

# Show all rows where any of col80-87 is non-null
highlighted = []
for r in all_rows:
    has_data = any(r[f'col{c}'] is not None for c in [80,81,82,83,84,85,86,87])
    if has_data:
        highlighted.append(r)
        b = str(r['branch'])[:24] if r['branch'] else ''
        base = f"{float(r['current_base']):,.0f}" if r['current_base'] else ''
        cps = f"{float(r['current_cps']):.4f}" if r['current_cps'] else ''
        c80 = f"{float(r['col80']):,.0f}" if r['col80'] is not None else ''
        c81 = f"{float(r['col81']):,.2f}" if r['col81'] is not None else ''
        c82 = f"{float(r['col82']):,.2f}" if r['col82'] is not None else ''
        c83 = f"{float(r['col83']):,.2f}" if r['col83'] is not None else ''
        c84 = f"{float(r['col84']):,.4f}" if r['col84'] is not None else ''
        c85 = f"{float(r['col85']):,.0f}" if r['col85'] is not None else ''
        print(f"{b:<25} {base:>10} {cps:>8} {c80:>12} {c81:>12} {c82:>12} {c83:>12} {c84:>12} {c85:>12}", flush=True)

print(f"\nRows with data in cols 80-87: {len(highlighted)}", flush=True)

# Also check actual column count
print(f"\nFirst row total column count: {len(list(ws.iter_rows(min_row=5, max_row=5))[0])}", flush=True)
