import openpyxl
import sys

FILE = r"C:\Users\chead\.openclaw\workspace\RBS_unzipped\RBS Budget Impact CCH.xlsx"
wb = openpyxl.load_workbook(FILE, read_only=True, data_only=True)

def read_sheet_full(ws, max_rows=200, max_cols=30):
    rows = []
    for row in ws.iter_rows(max_row=max_rows, max_col=max_cols, values_only=True):
        if any(v is not None for v in row):
            rows.append(row)
    return rows

# === SUMMARY TAB ===
print("=== SUMMARY TAB ===", flush=True)
ws_sum = wb["Summary"]
for row in ws_sum.iter_rows(max_row=50, max_col=25, values_only=True):
    if any(v is not None for v in row):
        vals = [str(v)[:30] if v is not None else '' for v in row]
        print(vals, flush=True)

# === BY TERRITORY - get totals (look for sum rows) ===
print("\n=== BY TERRITORY - Column Headers + Sample + Totals ===", flush=True)
ws_bt = wb["By Territory"]
row_count = 0
total_rows = []
for i, row in enumerate(ws_bt.iter_rows(max_col=25, values_only=True), 1):
    if any(v is not None for v in row):
        row_count += 1
        row_str = [str(v)[:25] if v is not None else '' for v in row]
        # Print header rows (1-3)
        if i <= 3:
            print(f"Row {i}: {row_str}", flush=True)
        # Look for total/summary rows
        if row[0] and str(row[0]).lower() in ['total', 'grand total', 'sum', 'company total']:
            print(f"TOTAL Row {i}: {row_str}", flush=True)
            total_rows.append((i, row))

print(f"Total data rows: {row_count}", flush=True)

# === BY FACILITY ===
print("\n=== BY FACILITY - First 20 rows ===", flush=True)
ws_fac = wb["By Facility"]
for i, row in enumerate(ws_fac.iter_rows(max_row=25, max_col=20, values_only=True), 1):
    if any(v is not None for v in row):
        vals = [str(v)[:25] if v is not None else '' for v in row]
        print(f"Row {i}: {vals}", flush=True)

# === COST INPUTS ===
print("\n=== COST INPUTS - First 20 rows ===", flush=True)
ws_ci = wb["Cost Inputs"]
for i, row in enumerate(ws_ci.iter_rows(max_row=25, max_col=20, values_only=True), 1):
    if any(v is not None for v in row):
        vals = [str(v)[:25] if v is not None else '' for v in row]
        print(f"Row {i}: {vals}", flush=True)

wb.close()
print("\nDONE", flush=True)
