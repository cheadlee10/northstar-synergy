import openpyxl
import json
from collections import defaultdict

FILE = r"C:\Users\chead\.openclaw\workspace\RBS_unzipped\RBS Budget Impact CCH.xlsx"
wb = openpyxl.load_workbook(FILE, read_only=True, data_only=True)

data = {}

# === SUMMARY ===
ws = wb["Summary"]
rows = []
for row in ws.iter_rows(values_only=True):
    if any(v is not None for v in row):
        rows.append([v for v in row])

# Full rollout
full_facs = rows[1][1]  # 95
full_savings = rows[2][2]  # 21.47M
full_vol = rows[2][3]  # 213.8M
full_cpp = rows[2][4]  # 0.100

# Facs with savings only
sav_facs = rows[5][1]  # 69
sav_savings = rows[6][2]  # 32.75M
sav_vol = rows[6][3]  # 145M
sav_cpp = rows[6][4]  # 0.226

# High savings low volume
hslow_count = rows[9][2]  # 16
hslow_impact = rows[9][3]  # 9.5M
hslow_vol = rows[9][4]  # 74K
hslow_cpp = rows[10][3]  # 0.29

# Savings w high error rates
herr_count = rows[13][2]  # 9
herr_impact = rows[13][3]  # 7.35M
herr_vol = rows[13][4]  # 110K
herr_cpp = rows[14][3]  # 0.22

# Forecast assumptions
asset_util = rows[1][7]  # 0.85
spr = rows[2][7]  # 1 (shipments per route)

data['summary'] = {
    'full_rollout': {'facs': full_facs, 'savings_2h': full_savings, 'volume_2h': full_vol, 'cpp_impact': full_cpp},
    'with_savings': {'facs': sav_facs, 'savings_2h': sav_savings, 'volume_2h': sav_vol, 'cpp_impact': sav_cpp},
    'high_sav_low_vol': {'count': hslow_count, 'impact': hslow_impact, 'avg_daily_vol': hslow_vol, 'cpp_impact': hslow_cpp},
    'high_error_rates': {'count': herr_count, 'impact': herr_impact, 'avg_daily_vol': herr_vol, 'cpp_impact': herr_cpp},
    'forecast_assumptions': {'asset_util': asset_util, 'shipments_per_route': spr}
}

# === BY FACILITY ===
ws = wb["By Facility"]
fac_rows = []
for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    if i > 4 and any(v is not None for v in row):
        fac_rows.append(row)

# Sort by adhoc % desc (column index 2)
top_adhoc = sorted(fac_rows, key=lambda r: float(r[2]) if r[2] is not None else 0, reverse=True)[:15]
# Sort by base desc
top_base = sorted(fac_rows, key=lambda r: float(r[3]) if r[3] is not None else 0, reverse=True)[:15]

data['by_facility'] = {
    'total_facilities': len(fac_rows),
    'top_adhoc_pct': [{'branch': r[1], 'adhoc_pct': float(r[2])*100, 'current_base': float(r[3]) if r[3] else 0, 'current_cps': float(r[4]) if r[4] else 0} for r in top_adhoc],
    'top_by_base': [{'branch': r[1], 'adhoc_pct': float(r[2])*100 if r[2] else 0, 'current_base': float(r[3]) if r[3] else 0, 'current_cps': float(r[4]) if r[4] else 0} for r in top_base]
}

# === LOUISVILLE ===
ws = wb["Louisville"]
lou_data = []
for i, row in enumerate(ws.iter_rows(max_row=20, values_only=True), 1):
    if any(v is not None for v in row):
        lou_data.append((i, list(row[:15])))

data['louisville'] = {
    'raw_rows': lou_data
}

# === COST INPUTS - sample ===
ws = wb["Cost Inputs"]
cost_rows = []
for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    if i > 1 and row[0] is not None:
        cost_rows.append({'fac': row[0], 'forecast_ppr': float(row[1]) if row[1] else 0, 'gas_maint_ppm': float(row[2]) if row[2] else 0, 'labor_ppd': float(row[3]) if row[3] else 0, 'avg_miles': float(row[4]) if row[4] else 0})

data['cost_inputs'] = {'total_facilities': len(cost_rows), 'sample': cost_rows[:5]}

wb.close()

# Output JSON
print(json.dumps(data, indent=2, default=str), flush=True)
