import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import smtplib, os
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

CHART_DIR = r"C:\Users\chead\.openclaw\workspace\chart_output"
OUT       = r"C:\Users\chead\OneDrive\Documents\RBS_Savings_Analysis.xlsx"

# Load existing workbook and replace Charts sheet
wb = openpyxl.load_workbook(OUT)
if "Charts" in wb.sheetnames:
    del wb["Charts"]

NAVY  = "1F3864"
MID   = "2F5597"
LGREY = "F2F2F2"
thin  = Side(style='thin', color="BFBFBF")
bdr   = Border(left=thin, right=thin, bottom=thin, top=thin)

ws = wb.create_sheet("Charts")
wb.move_sheet("Charts", offset=-(len(wb.sheetnames)-1))  # move to front
ws.sheet_view.showGridLines = False
ws.sheet_tab_color = MID

# Title
ws.merge_cells('B1:T1')
ws['B1'] = "RBS Pilot — Savings Driver Charts"
ws['B1'].font = Font(name="Calibri", bold=True, size=15, color=NAVY)
ws['B1'].alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 30

ws.merge_cells('B2:T2')
ws['B2'] = "Three charts telling the complete story: HOW MUCH each facility saves  ·  WHY they save  ·  TWO PATHS to savings"
ws['B2'].font = Font(name="Calibri", size=10, color="595959", italic=True)
ws['B2'].alignment = Alignment(horizontal="left")
ws.row_dimensions[2].height = 16
ws.row_dimensions[3].height = 8

# Chart 1 — CPP Comparison (most important, full width at top)
ws.merge_cells('B4:T4')
ws['B4'] = "Chart 1 — RBS vs. Stop Payment CPP by Facility  (sorted: largest savings opportunity at top)"
ws['B4'].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
ws['B4'].fill = PatternFill("solid", fgColor=NAVY)
ws['B4'].alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[4].height = 22

img1 = XLImage(os.path.join(CHART_DIR, 'chart1_cpp_comparison.png'))
img1.anchor = 'B5'
img1.width  = 900
img1.height = 530
ws.add_image(img1)

# Row spacer after chart 1 (approx 30 rows for the image height)
for r in range(5, 38):
    ws.row_dimensions[r].height = 18

ws.row_dimensions[38].height = 14

# Chart 2 — Density driver
ws.merge_cells('B39:T39')
ws['B39'] = "Chart 2 — What Drives Savings? Route Density vs. CPP Gap  (bubble size = annual savings)"
ws['B39'].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
ws['B39'].fill = PatternFill("solid", fgColor=MID)
ws['B39'].alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[39].height = 22

img2 = XLImage(os.path.join(CHART_DIR, 'chart2_density_driver.png'))
img2.anchor = 'B40'
img2.width  = 900
img2.height = 560
ws.add_image(img2)

for r in range(40, 75):
    ws.row_dimensions[r].height = 18

ws.row_dimensions[75].height = 14

# Chart 3 — Two Paths
ws.merge_cells('B76:T76')
ws['B76'] = "Chart 3 — Two Paths to Savings: High Volume vs. High CPP Gap  (point size = CPP gap magnitude)"
ws['B76'].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
ws['B76'].fill = PatternFill("solid", fgColor=NAVY)
ws['B76'].alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[76].height = 22

img3 = XLImage(os.path.join(CHART_DIR, 'chart3_volume_vs_savings.png'))
img3.anchor = 'B77'
img3.width  = 900
img3.height = 560
ws.add_image(img3)

wb.save(OUT)
print(f"Saved: {OUT}")

# ── EMAIL ──────────────────────────────────────────────────────────────────────
FROM_EMAIL   = "clawcliff@gmail.com"
APP_PASSWORD = "zpon bjsp dnfx tkdy"
TO_EMAIL     = "C963146@ontrac.com"

HTML = """<html><body style="font-family:Calibri,Arial,sans-serif;font-size:14px;color:#1F3864;">
<p>Craig,</p>
<p>Updated with three analytical charts that tell the full story of what's driving the RBS savings.</p>

<h3 style="color:#2F5597;">The Three Charts</h3>
<ol>
<li><b>RBS vs. Stop Payment CPP by Facility</b> — Side-by-side comparison showing exactly how far the CPP drops at each site. Sorted so the largest opportunity (Toledo, $1.15 gap) is at the top. Delta and annual savings annotated inline.</li>
<li><b>Route Density vs. CPP Gap (bubble chart)</b> — The analytical driver chart. Lower density (packages/route) = bigger CPP gap = bigger RBS advantage. Toledo is flagged as an outlier: low volume + unusually high stop fee rate. Trend line shows the relationship across the other 11.</li>
<li><b>Two Paths to Savings (scatter)</b> — High-volume facilities (Dallas, Northern NJ, Atlanta) save through scale even with a modest CPP gap. Low-volume facilities (Toledo, Boise, Pittsburgh) save through an outsized CPP gap. Both paths contribute to the $11.35M total.</li>
</ol>

<p><b>Total Annual Savings (12 Highlighted Facilities): $11,353,251</b></p>
<p style="color:#595959;font-size:12px;margin-top:24px;">Cliff | Financial Analysis</p>
</body></html>"""

msg = MIMEMultipart("alternative")
msg["From"]    = f"Cliff <{FROM_EMAIL}>"
msg["To"]      = TO_EMAIL
msg["Subject"] = "RBS Pilot — Savings Analysis with Charts | 12 Facilities"
msg.attach(MIMEText(HTML, "html"))

with open(OUT, "rb") as f:
    part = MIMEBase("application","vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    part.set_payload(f.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition",'attachment; filename="RBS_Savings_Analysis.xlsx"')
    msg.attach(part)

with smtplib.SMTP("smtp.gmail.com", 587) as s:
    s.starttls()
    s.login(FROM_EMAIL, APP_PASSWORD)
    s.sendmail(FROM_EMAIL, TO_EMAIL, msg.as_string())

print("Email sent.")
