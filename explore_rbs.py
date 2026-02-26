import openpyxl
import sys

FILE = r"C:\Users\chead\.openclaw\workspace\RBS_unzipped\RBS Budget Impact CCH.xlsx"

# Read-only mode - much lower memory
wb = openpyxl.load_workbook(FILE, read_only=True, data_only=True)

print(f"Sheet count: {len(wb.sheetnames)}", flush=True)
for name in wb.sheetnames:
    print(f"  Sheet: {name}", flush=True)

# Read first sheet
ws = wb.worksheets[0]
print(f"\n=== First sheet: '{ws.title}' ===", flush=True)
rows_read = 0
for row in ws.iter_rows(max_row=15, values_only=True):
    vals = [str(v)[:25] if v is not None else '' for v in row[:20]]
    print(f"  {vals}", flush=True)
    rows_read += 1

wb.close()
print("DONE", flush=True)
